"""
Phase 2 — Full Insertion
All 46 class names are fully resolved (hardcoded from Phase 1).
No new classtypes needed — all 35 were already created.
"""
import sys
import subprocess

sys.stdout.reconfigure(encoding="utf-8", errors="replace")

for pkg in ("openpyxl", "python-dotenv", "supabase"):
    try:
        __import__(pkg.replace("-", "_").split("[")[0])
    except ImportError:
        print(f"Installing {pkg}...")
        subprocess.check_call([sys.executable, "-m", "pip", "install", pkg, "-q"])

import re
from pathlib import Path
from datetime import datetime, date, time, timedelta
from zoneinfo import ZoneInfo
from dotenv import load_dotenv
import os
import openpyxl

load_dotenv(Path(r"C:\Users\oren ahuvia\Documents\GitHub\RideOn-Project\.env"))
from supabase import create_client
sb = create_client(os.environ["SUPABASE_URL"], os.environ["SUPABASE_KEY"])

# ── Constants ────────────────────────────────────────────────────────────────
HOST_RANCH_ID = 11
ARENA_ID      = 1
CREATOR_UID   = 2
FIELD_ID      = 1
TZ            = ZoneInfo("Asia/Jerusalem")
EXCEL         = Path(r"C:\Users\oren ahuvia\Documents\GitHub\RideOn-Project\סיכום תחרויות ריינינג 25.xlsx")
ENTRY_START   = 10000

# ── Complete classname → classtypeid map (all 46 Excel names, lower-stripped) ─
CLASSTYPE_MAP = {
    # original exact matches
    "ירוקי התאחדות":                       2,
    "ירוקי רוכב חדש התאחדות":              3,
    "נוביס התאחדות":                       10,
    "נוביס נונ פרו התאחדות":               11,
    "נונ פרו 50+":                          6,
    "נוער ירוקי התאחדות":                   4,
    "פתוח לא מוגבל":                       1,
    # fuzzy-corrected (use real Reining classes)
    "prime time non-professional riders":  14,
    "נוער 14-18 nrha":                     15,
    "פתוח לא מוגבל b2w":                   1,
    "פתוח לא מוגבל dk":                    1,
    # new classtypes created in previous step (ids 123-157)
    "irbc דרבי נונ פרו l1":               123,
    "irbc דרבי נונ פרו l3":               124,
    "irbc דרבי נונ פרו l4":               125,
    "irbc דרבי נונ פרו פריים טיים":        126,
    "irbc דרבי פתוח l1":                  127,
    "irbc דרבי פתוח l3":                  128,
    "irbc דרבי פתוח l4":                  129,
    "junior riders":                       130,
    "senior non-professional riders":      131,
    "senior professional riders":          132,
    "young non-professional riders":       133,
    "young professional riders":           134,
    "אקסטרים שוטאווט":                    135,
    "דרבי נונ פרו l4":                    136,
    "דרבי פתוח l4":                       137,
    "זוגות 1":                             138,
    "זוגות 2":                             139,
    "נוביס l1 nrha":                      140,
    "נונ פרו nrha":                        141,
    "נונ פרו מוגבל nrha":                  142,
    "נונ פרו פריים טיים 40+ nrha":         143,
    "נוער 13 nrha":                        144,
    "נוער unrestricted nrha":              145,
    "פטוריטי נונ פרו nrha":               146,
    "פטוריטי נונ פרו בני 4 irbc l1":      147,
    "פטוריטי נונ פרו בני 4 irbc l2":      148,
    "פטוריטי נונ פרו בני 4 irbc l3":      149,
    "פטוריטי נונ פרו בני 4 irbc l4":      150,
    "פטוריטי פתוח nrha":                  151,
    "פטוריטי פתוח בני 3 irbc l1":         152,
    "פטוריטי פתוח בני 3 irbc l2":         153,
    "פטוריטי פתוח בני 3 irbc l3":         154,
    "פטוריטי פתוח בני 3 irbc l4":         155,
    "פתוח nrha":                           156,
    "פתוח מוגבל nrha":                    157,
}

def resolve(name):
    return CLASSTYPE_MAP.get(str(name).strip().lower())

# ── Skip / day-boundary helpers ──────────────────────────────────────────────
SKIP_PREFIXES = ["פייד טיים", "מספר", "מקצה לרישום דמי ביטול", 'סה"כ']

def should_skip(name):
    if not name:
        return True
    n = str(name).strip()
    if not n or n == "מספר":
        return True
    for p in SKIP_PREFIXES:
        if n.startswith(p):
            return True
    try:
        float(n)
        return True
    except ValueError:
        pass
    return False

def is_paid_time(name):
    return bool(name) and str(name).strip().startswith("פייד טיים")

DAY_OFFSETS = {"שני": 1, "שלישי": 2, "רביעי": 3, "חמישי": 4, "שישי": 5}

def day_offset(text):
    for day, off in DAY_OFFSETS.items():
        if day in str(text):
            return off
    return 0

def parse_dates(cell):
    s = str(cell).strip()
    m = re.search(r"(\d+)-(\d+)\.(\d+)\.(\d+)", s)
    if m:
        return date(int(m[4]), int(m[3]), int(m[1])), date(int(m[4]), int(m[3]), int(m[2]))
    m2 = re.search(r"(\d+)\.(\d+)\.(\d+)", s)
    if m2:
        d = date(int(m2[3]), int(m2[2]), int(m2[1]))
        return d, d
    raise ValueError(f"Cannot parse date: {cell!r}")

def to_float(v):
    try:
        return float(v) if v not in (None, "") else 0.0
    except Exception:
        return 0.0

# ── Stats ────────────────────────────────────────────────────────────────────
S = dict(comps_ins=0, comps_skip=0, cls_ins=0, cls_skip=0,
         prizes=0, entries=0, rows_skip=0, errors=0)

# ── Load Excel ───────────────────────────────────────────────────────────────
print(f"Loading Excel: {EXCEL.name}")
wb        = openpyxl.load_workbook(EXCEL, data_only=True)
names     = wb.sheetnames
gilyonot  = [s for s in names if s.startswith("גיליון")]
summaries = [s for s in names if not s.startswith("גיליון")]
print(f"  {len(gilyonot)} class sheets, {len(summaries)} summary sheets\n")

entries_todo = []   # (classincompid, entrycount, classdatetime, competitionid)

# ════════════════════════════════════════════════════════════════════════════
# STEP 1+2  Competitions + classincompetition
# ════════════════════════════════════════════════════════════════════════════
print("STEP 1+2 — Competitions and Classes")
print("─" * 62)

for sheet in gilyonot:
    ws   = wb[sheet]
    rows = list(ws.iter_rows(values_only=True))

    comp_name              = str(rows[0][0]).strip() if rows[0][0] else sheet
    event_start, event_end = parse_dates(rows[1][0])
    print(f"\n  [{sheet}]  {comp_name}  ({event_start} – {event_end})")

    # Step 1: insert competition (idempotent)
    ex = sb.table("competition").select("competitionid").eq("competitionname", comp_name).execute()
    if ex.data:
        comp_id = ex.data[0]["competitionid"]
        print(f"    SKIP competition (exists id={comp_id})")
        S["comps_skip"] += 1
    else:
        try:
            r = sb.table("competition").insert({
                "hostranchid":           HOST_RANCH_ID,
                "fieldid":               FIELD_ID,
                "createdbysystemuserid": CREATOR_UID,
                "competitionname":       comp_name,
                "competitionstartdate":  event_start.isoformat(),
                "competitionenddate":    event_end.isoformat(),
                "competitionstatus":     "הסתיימה",
            }).execute()
            comp_id = r.data[0]["competitionid"]
            print(f"    INSERTED competition id={comp_id}")
            S["comps_ins"] += 1
        except Exception as e:
            print(f"    ERROR competition: {e}")
            S["errors"] += 1
            continue

    # Resolve header columns
    hdr = rows[2]
    ci_ent = ci_org = ci_fed = None
    for i, h in enumerate(hdr):
        if h is None:
            continue
        hs = str(h).strip()
        if "מספר כניסות" in hs:
            ci_ent = i
        elif hs == "חווה":
            ci_org = i
        elif hs == "התאחדות":
            ci_fed = i

    # Step 2: classincompetition
    cur_off   = 0
    day_order = {}

    for row in rows[3:]:
        if row[0] is None:
            continue
        raw = str(row[0]).strip()

        if is_paid_time(raw):
            cur_off = day_offset(raw)
            continue

        if should_skip(raw):
            S["rows_skip"] += 1
            continue

        ctype_id = resolve(raw)
        if ctype_id is None:
            print(f"    WARNING: no mapping for {raw!r}")
            S["rows_skip"] += 1
            continue

        n_entries = int(to_float(row[ci_ent])) if ci_ent is not None else 0
        org_cost  = to_float(row[ci_org])       if ci_org  is not None else 0.0
        fed_cost  = to_float(row[ci_fed])       if ci_fed  is not None else 0.0

        cls_date  = event_start + timedelta(days=cur_off)
        cls_dt    = datetime.combine(cls_date, time(9, 0)).replace(tzinfo=TZ)
        date_iso  = cls_date.isoformat()

        day_order[cur_off] = day_order.get(cur_off, 0) + 1
        order = day_order[cur_off]

        # Idempotency: (competitionid, classtypeid, date)
        ex_cls = (
            sb.table("classincompetition").select("classincompid")
            .eq("competitionid", comp_id)
            .eq("classtypeid",   ctype_id)
            .gte("classdatetime", f"{date_iso}T00:00:00+00:00")
            .lte("classdatetime", f"{date_iso}T23:59:59+00:00")
            .execute()
        )
        if ex_cls.data:
            cic_id = ex_cls.data[0]["classincompid"]
            S["cls_skip"] += 1
            if n_entries > 0:
                ex_e = sb.table("entry").select("entryid").eq("classincompid", cic_id).execute()
                if len(ex_e.data) < n_entries:
                    entries_todo.append((cic_id, n_entries, cls_dt, comp_id))
            continue

        try:
            r = sb.table("classincompetition").insert({
                "competitionid":  comp_id,
                "classtypeid":    ctype_id,
                "arenaranchid":   HOST_RANCH_ID,
                "arenaid":        ARENA_ID,
                "classdatetime":  cls_dt.isoformat(),
                "organizercost":  org_cost,
                "federationcost": fed_cost,
                "orderinday":     order,
            }).execute()
            cic_id = r.data[0]["classincompid"]
            S["cls_ins"] += 1
            if n_entries > 0:
                entries_todo.append((cic_id, n_entries, cls_dt, comp_id))
        except Exception as e:
            print(f"    ERROR class {raw!r}: {e}")
            S["errors"] += 1

print(f"\n  Competitions: {S['comps_ins']} inserted, {S['comps_skip']} skipped")
print(f"  Classes:      {S['cls_ins']} inserted, {S['cls_skip']} skipped")

# ════════════════════════════════════════════════════════════════════════════
# STEP 3  Prizes
# ════════════════════════════════════════════════════════════════════════════
print("\nSTEP 3 — Prizes")
print("─" * 62)

for gil_sheet, sum_sheet in zip(gilyonot, summaries):
    ws_g      = wb[gil_sheet]
    comp_name = str(list(ws_g.iter_rows(values_only=True))[0][0]).strip()
    comp_r    = sb.table("competition").select("competitionid").eq("competitionname", comp_name).execute()
    if not comp_r.data:
        print(f"  WARNING: competition not found: {comp_name!r}")
        continue
    comp_id = comp_r.data[0]["competitionid"]

    ws_s  = wb[sum_sheet]
    srows = list(ws_s.iter_rows(values_only=True))
    hdr   = srows[2]

    ci_name = ci_jackpot = ci_fixed = None
    for i, h in enumerate(hdr):
        if h is None:
            continue
        hs = str(h).strip()
        if hs == "שם המקצה":
            ci_name = i
        elif hs == "תשלום לפרס":
            ci_jackpot = i
        elif "פרס כספי" in hs:
            ci_fixed = i

    if ci_name is None:
        print(f"  WARNING: no 'שם המקצה' column in {sum_sheet}")
        continue

    for row in srows[3:]:
        if len(row) <= ci_name or row[ci_name] is None:
            continue
        raw = str(row[ci_name]).strip()
        if not raw or raw.startswith("סה"):
            continue

        ctype_id = resolve(raw)
        if ctype_id is None:
            continue

        cic_r = (
            sb.table("classincompetition").select("classincompid")
            .eq("competitionid", comp_id)
            .eq("classtypeid",   ctype_id)
            .execute()
        )
        if not cic_r.data:
            continue

        jackpot = to_float(row[ci_jackpot]) if ci_jackpot is not None and len(row) > ci_jackpot else 0.0
        fixed   = to_float(row[ci_fixed])   if ci_fixed   is not None and len(row) > ci_fixed   else 0.0

        for cic_row in cic_r.data:
            cic_id = cic_row["classincompid"]
            if jackpot > 0:
                try:
                    if not sb.table("classprize").select("classincompid").eq("classincompid", cic_id).eq("prizetypeid", 2).execute().data:
                        sb.table("classprize").insert({"classincompid": cic_id, "prizetypeid": 2, "prizeamount": jackpot}).execute()
                        S["prizes"] += 1
                except Exception as e:
                    print(f"  ERROR prize jackpot cic={cic_id}: {e}")
                    S["errors"] += 1
            if fixed > 0:
                try:
                    if not sb.table("classprize").select("classincompid").eq("classincompid", cic_id).eq("prizetypeid", 3).execute().data:
                        sb.table("classprize").insert({"classincompid": cic_id, "prizetypeid": 3, "prizeamount": fixed}).execute()
                        S["prizes"] += 1
                except Exception as e:
                    print(f"  ERROR prize fixed cic={cic_id}: {e}")
                    S["errors"] += 1

print(f"  Prizes inserted: {S['prizes']}")

# ════════════════════════════════════════════════════════════════════════════
# STEP 4  Fabricate entries
# ════════════════════════════════════════════════════════════════════════════
print(f"\nSTEP 4 — Fabricating entries ({len(entries_todo)} classes to process)")
print("─" * 62)

counter = ENTRY_START

for cic_id, n_entries, cls_dt, comp_id in entries_todo:
    ex_e    = sb.table("entry").select("entryid").eq("classincompid", cic_id).execute()
    already = len(ex_e.data)

    if already >= n_entries:
        print(f"  SKIP cic={cic_id} ({already}/{n_entries} entries already exist)")
        counter += n_entries
        continue

    start_n   = counter + already
    to_create = n_entries - already

    for i in range(to_create):
        n = start_n + i
        try:
            p = sb.table("person").insert({
                "nationalid": f"{n:09d}",
                "firstname":  "רוכב",
                "lastname":   f"היסטורי {n}",
            }).execute().data[0]
            pid = p["personid"]

            sb.table("federationmember").insert({
                "federationmemberid": pid,
                "hasvalidmembership": True,
            }).execute()

            h = sb.table("horse").insert({
                "ranchid":   HOST_RANCH_ID,
                "horsename": f"סוס היסטורי {n}",
            }).execute().data[0]

            sb.table("systemuser").insert({
                "systemuserid": pid,
                "username":     f"hist_{n}",
                "passwordhash": "placeholder",
                "passwordsalt": "placeholder",
                "isactive":     False,
            }).execute()

            bill = sb.table("bill").insert({
                "paidbypersonid": pid,
                "amounttopay":    0,
                "dateopened":     cls_dt.isoformat(),
                "competitionid":  comp_id,
            }).execute().data[0]

            sr = sb.table("servicerequest").insert({
                "orderedbysystemuserid":   pid,
                "horseid":                 h["horseid"],
                "riderfederationmemberid": pid,
                "billid":                  bill["billid"],
            }).execute().data[0]

            sb.table("entry").insert({
                "entryid":      sr["srequestid"],
                "classincompid": cic_id,
                "entrystatus":  "Active",
            }).execute()

            S["entries"] += 1

        except Exception as e:
            print(f"  ERROR entry n={n} cic={cic_id}: {e}")
            S["errors"] += 1

        if S["entries"] > 0 and S["entries"] % 50 == 0:
            print(f"  ... {S['entries']} entries fabricated so far")

    counter += n_entries
    print(f"  cic={cic_id}: {to_create} entries fabricated")

# ════════════════════════════════════════════════════════════════════════════
# SUMMARY
# ════════════════════════════════════════════════════════════════════════════
print()
print("=" * 62)
print("INSERTION COMPLETE")
print("=" * 62)
print(f"  Competitions inserted  : {S['comps_ins']}")
print(f"  Classes inserted       : {S['cls_ins']}")
print(f"  Prizes inserted        : {S['prizes']}")
print(f"  Entries fabricated     : {S['entries']}")
print(f"  Rows skipped           : {S['rows_skip']}")
print(f"  Errors                 : {S['errors']}")
print("=" * 62)
