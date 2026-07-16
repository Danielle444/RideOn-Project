"""
Phase 1 — Class Name Matching Report for All-Around 2025 (Format D).
Uses openpyxl for row-color-based day boundary detection.
"""
import sys
import subprocess

sys.stdout.reconfigure(encoding="utf-8", errors="replace")

for pkg in ("openpyxl", "pandas", "python-dotenv", "supabase"):
    try:
        __import__(pkg.replace("-", "_").split("[")[0])
    except ImportError:
        subprocess.check_call([sys.executable, "-m", "pip", "install", pkg, "-q"])

import re
import difflib
from pathlib import Path
from datetime import date
from dotenv import load_dotenv
import os
import pandas as pd
from openpyxl import load_workbook

load_dotenv(Path(r"C:\Users\oren ahuvia\Documents\GitHub\RideOn-Project\.env"))
from supabase import create_client
sb = create_client(os.environ["SUPABASE_URL"], os.environ["SUPABASE_KEY"])

EXCEL   = Path(r"C:\Users\oren ahuvia\Documents\GitHub\RideOn-Project\תחרויות אולארונד 2025.xlsx")
PDF_DIR = Path(r"C:\Users\oren ahuvia\Documents\GitHub\RideOn-Project\pdfs_allround")

print(f"Loading: {EXCEL.name}\n")

# ── Color → day index mapping (same across all sheets) ────────────────────────
COLOR_TO_DAY = {
    "FFD9E1F2": 0,   # blue   = first color group in most sheets
    "FFFCE4D6": 1,   # orange = second
    "FFFFF2CC": 2,   # yellow = third
}
# Winter sheet has reversed order (yellow=day0, blue=day1, orange=day2)
WINTER_COLOR_TO_DAY = {
    "FFFFF2CC": 0,   # yellow = day 1 (Dec 11)
    "FFD9E1F2": 1,   # blue   = day 2 (Dec 12)
    "FFFCE4D6": 2,   # orange = day 3 (Dec 13)
}

def get_row_color(ws, row_idx):
    for cell in ws[row_idx]:
        if cell.fill and cell.fill.fgColor:
            fc = cell.fill.fgColor
            if fc.type == "rgb" and fc.rgb not in ("00000000", "FF000000"):
                return fc.rgb
    return None

def parse_date_range(s):
    m = re.search(r"(\d{1,2})[-–](\d{1,2})[./](\d{1,2})[./](\d{2,4})", str(s))
    if not m:
        return None, None
    d1, d2, mo, yr = int(m[1]), int(m[2]), int(m[3]), int(m[4])
    if yr < 100:
        yr += 2000
    return date(yr, mo, d1), date(yr, mo, d2)

DATE_SPLIT_RE = re.compile(r"(\d{1,2})[-–](\d{1,2})[./](\d{1,2})[./](\d{2,4})")

def all_dates_in_range(start_d, end_d):
    days = []
    cur = start_d
    while cur <= end_d:
        days.append(cur)
        from datetime import timedelta
        cur += timedelta(days=1)
    return days

# ── Class name normalization ──────────────────────────────────────────────────
TYPO_CORRECTIONS = {
    "הורסמני פתוח":                    "הורסמנשיפ פתוח",
    "האנט סיט אקויטיישן עד גיל 13 ירוקי": "האנט סיט אקוויטיישן עד גיל 13 ירוקי",
    "האנט סיט אקויטיישן":               "האנט סיט אקוויטיישן",
    "הורסמשניפ שוטאאוט 20":            "הורסמנשיפ שוטאאוט",
    "טרייל שוטאאוט 20":                "טרייל שוטאאוט",
    "פלז'ר שוט אאוט":                  "פלז'ר שואוטאוט",
}

def normalize(raw):
    if not isinstance(raw, str):
        return ""
    n = raw.strip()
    # Apply typo corrections
    for wrong, right in TYPO_CORRECTIONS.items():
        if n == wrong or n.lower() == wrong.lower():
            n = right
            break
    # Collapse multiple spaces
    n = re.sub(r"\s+", " ", n).strip()
    return n

# ── Skip rules ────────────────────────────────────────────────────────────────
SKIP_STARTS = ["מקצה לרישום דמי ביטול", "מקצה לביטול", 'סה"כ', "סה\"כ"]
def should_skip(raw):
    if not raw or str(raw).strip() in ("", "nan"):
        return True
    n = str(raw).strip()
    if not n or n == "מספר":
        return True
    for s in SKIP_STARTS:
        if n.startswith(s):
            return True
    try:
        float(n)
        return True
    except ValueError:
        pass
    return False

# ── Fetch existing classtypes ─────────────────────────────────────────────────
print("Fetching classtypes from Supabase...")
existing    = sb.table("classtype").select("classtypeid, classname, fieldid").execute().data
print(f"  {len(existing)} total classtypes\n")
ex_names    = {r["classname"]: r["classtypeid"] for r in existing}
ex_lower    = {}
for r in existing:
    k = re.sub(r"\s+", " ", r["classname"]).strip().lower()
    ex_lower[k] = (r["classname"], r["classtypeid"])
ex_keys     = list(ex_names.keys())

# ── Additional known mappings (user-confirmed) ─────────────────────────────────
USER_MAPPINGS = {
    "פלז'ר נונ פרו 40+ הליכה ג'וג":        95,
    "הורסמנשיפ נונ פרו 40+ הליכה ג'וג":    112,
}

# ── Load workbook for colors, pandas for values ───────────────────────────────
wb          = load_workbook(EXCEL, data_only=True)
all_dfs     = pd.read_excel(EXCEL, sheet_name=None, header=None, dtype=str)

# ── Split-detection for mixed sheets (Cutting block) ─────────────────────────
CUTTING_SPLIT_RE = re.compile(r"קאטינג\s+\d+\s+\d{1,2}[-–]\d{1,2}\.\d{2}")

# ── Main analysis ─────────────────────────────────────────────────────────────
all_unique  = {}   # normalized_lower → (raw_first, sheet_name, fieldid)
comp_blocks = []   # list of competition info dicts

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    df = all_dfs[sheet_name]

    # Competition name and dates from rows 0-1
    comp_name  = str(df.iloc[0, 0]).strip() if pd.notna(df.iloc[0, 0]) else sheet_name
    date_raw   = str(df.iloc[1, 0]).strip() if len(df) > 1 and pd.notna(df.iloc[1, 0]) else ""
    start_d, end_d = parse_date_range(date_raw)
    days_list  = all_dates_in_range(start_d, end_d) if start_d else []

    is_winter  = "חורף" in sheet_name
    color_map  = WINTER_COLOR_TO_DAY if is_winter else COLOR_TO_DAY
    field_id   = 3   # All-Around by default

    # Find where Cutting block starts
    cutting_start = None
    for row_idx in range(4, len(df)):
        raw_val = str(df.iloc[row_idx, 0]).strip() if pd.notna(df.iloc[row_idx, 0]) else ""
        if CUTTING_SPLIT_RE.match(raw_val):
            cutting_start = row_idx
            break

    # Color groups for All-Around block
    color_groups_detected = []
    prev_color = "__"
    for row in ws.iter_rows(min_row=4):
        r_idx = row[0].row
        # Stop at cutting split
        if cutting_start and r_idx > cutting_start + 2:
            break
        raw_val = str(row[0].value or "").strip()
        if not raw_val:
            continue
        color = get_row_color(ws, r_idx)
        if color != prev_color:
            if color in color_map:
                color_groups_detected.append((color, color_map.get(color, -1)))
            prev_color = color

    # Unique color groups in order
    seen_colors = []
    for c, idx in color_groups_detected:
        if c not in [x[0] for x in seen_colors]:
            seen_colors.append((c, idx))

    comp_blocks.append({
        "sheet": sheet_name, "comp_name": comp_name,
        "start": start_d, "end": end_d, "days": days_list,
        "field_id": 3, "color_map": color_map,
        "color_groups": seen_colors,
        "cutting_split": cutting_start,
    })

    # Collect class names from All-Around block
    for row in ws.iter_rows(min_row=4):
        r_idx = row[0].row
        if cutting_start and r_idx >= cutting_start:
            break   # stop before Cutting block
        raw_val = str(row[0].value or "").strip()
        if not raw_val or should_skip(raw_val):
            continue
        norm = normalize(raw_val)
        if not norm:
            continue
        key = re.sub(r"\s+", " ", norm).lower()
        if key not in all_unique:
            all_unique[key] = (raw_val, sheet_name, 3)

    # Collect Cutting block class names
    if cutting_start:
        # Cutting header: row cutting_start = date row, cutting_start+1 = column headers
        cutting_comp_date = str(df.iloc[cutting_start, 0]).strip()

        for ri in range(cutting_start + 2, len(df)):
            raw_val = str(df.iloc[ri, 0]).strip() if pd.notna(df.iloc[ri, 0]) else ""
            if not raw_val or should_skip(raw_val):
                continue
            norm = normalize(raw_val)
            if not norm:
                continue
            key = re.sub(r"\s+", " ", norm).lower()
            if key not in all_unique:
                all_unique[key] = (raw_val, sheet_name + " [Cutting]", 2)

print(f"Unique normalized class names: {len(all_unique)}\n")

# ── Match classes ─────────────────────────────────────────────────────────────
exact_rows = []
fuzzy_rows = []
new_rows   = []

for key, (raw_first, sheet_first, fid) in sorted(all_unique.items(),
                                                   key=lambda x: (x[1][2], x[0])):
    norm = re.sub(r"\s+", " ", normalize(raw_first)).strip()
    if not norm:
        continue

    # User-specified mapping
    user_cid = USER_MAPPINGS.get(norm) or USER_MAPPINGS.get(raw_first.strip())
    if user_cid:
        db_r = next((r for r in existing if r["classtypeid"] == user_cid), {})
        exact_rows.append((raw_first, norm, fid, user_cid, db_r.get("classname", "?"), "USER-MAP"))
        continue

    norm_key = key  # already lower+collapsed

    # 1. Exact (original string)
    if norm in ex_names:
        exact_rows.append((raw_first, norm, fid, ex_names[norm], norm, "EXACT"))
        continue

    # 2. Case-insensitive + space-collapsed exact
    if norm_key in ex_lower:
        db_name, cid = ex_lower[norm_key]
        exact_rows.append((raw_first, norm, fid, cid, db_name, "EXACT-CI"))
        continue

    # 3. Fuzzy
    close = difflib.get_close_matches(norm, ex_keys, n=1, cutoff=0.55)
    if close:
        score = round(difflib.SequenceMatcher(None, norm, close[0]).ratio(), 2)
        fuzzy_rows.append((raw_first, norm, fid, ex_names[close[0]], close[0], score))
    else:
        new_rows.append((raw_first, norm, fid))

# ── Prize mapping (from skill specification) ──────────────────────────────────
PRIZE_JACKPOT_100 = {
    "טרייל פתוח", "טרייל נונ פרו", "טרייל פתוח לסוסי נוביס", "טרייל ירוקי בוגרים",
    "האנט סיט אקוויטיישן פתוח", "האנטר אנדר סאדל פתוח",
    "הורסמנשיפ פתוח", "הורסמנשיפ ירוקי בוגרים", "הורסמנשיפ נונ פרו",
    "הורסמנשיפ פתוח לסוסי נוביס",
    "פלז'ר פתוח", "פלז'ר נונ פרו", "פלז'ר ירוקי בוגרים", "פלז'ר פתוח לסוסי נוביס",
}
PRIZE_CIRCUIT = {  # jackpot ₪100 + כסף מוסף ₪2,000
    "הורסמנשיפ סניור - סירקט", "הורסמנשיפ ג'וניור - סירקט", "הורסמנשיפ סירקט",
    "טרייל סירקט", "האנטר אנדר סאדל סירקט",
    "פלז'ר סירקט סניור", "פלז'ר סירקט ג'וניור",
    "הליכה ג'וג סירקט עד 13",
    "שואומנשיפ סירקט",
}
PRIZE_SHOOTOUT = {  # כסף מוסף ₪8,000 only
    "טרייל שוטאאוט", "פלז'ר שואוטאוט", "הורסמנשיפ שוטאווט", "הורסמנשיפ שוטאאוט",
}

def prize_str(norm):
    if norm in PRIZE_SHOOTOUT:
        return "כסף מוסף ₪8,000"
    if norm in PRIZE_CIRCUIT:
        return "ג'קפוט ₪100 + כסף מוסף ₪2,000"
    if norm in PRIZE_JACKPOT_100:
        return "ג'קפוט ₪100"
    return ""

# ── Print report ──────────────────────────────────────────────────────────────
FIELD_LABEL = {3: "אולארונד", 2: "קאטינג"}

print("=" * 74)
print("CLASS NAME MATCHING REPORT — תחרויות אולארונד 2025")
print("=" * 74)

print(f"\n--- EXACT / USER MATCHES ({len(exact_rows)}) ---")
for raw, norm, fid, cid, db_name, method in sorted(exact_rows, key=lambda x: x[2]):
    typo = f"  (typo-corrected from '{raw}')" if norm != raw.strip() else ""
    pr   = prize_str(norm)
    pr_s = f"  [{pr}]" if pr else ""
    label = FIELD_LABEL.get(fid, "?")
    print(f"  [{label}]  '{norm}'{typo}")
    print(f"    {method} → id={cid}  db='{db_name}'{pr_s}")

print(f"\n--- FUZZY MATCHES ({len(fuzzy_rows)}) ---")
for raw, norm, fid, cid, db_name, score in sorted(fuzzy_rows, key=lambda x: (x[2], x[1])):
    typo = f"  (from '{raw}')" if norm != raw.strip() else ""
    label = FIELD_LABEL.get(fid, "?")
    print(f"  [{label}]  '{norm}'{typo}")
    print(f"    FUZZY → id={cid}  db='{db_name}'  score={score}")

print(f"\n--- NEW CLASSTYPES NEEDED — fieldid=3 ({sum(1 for _,_,f in new_rows if f==3)}) ---")
for raw, norm, fid in sorted(new_rows, key=lambda x: x[1]):
    if fid == 3:
        pr = prize_str(norm)
        pr_s = f"  [{pr}]" if pr else ""
        print(f"  [NEW-אולארונד]  '{norm}'{pr_s}")

print(f"\n--- NEW CLASSTYPES NEEDED — fieldid=2 ({sum(1 for _,_,f in new_rows if f==2)}) ---")
for raw, norm, fid in sorted(new_rows, key=lambda x: x[1]):
    if fid == 2:
        print(f"  [NEW-קאטינג]    '{norm}'")

print()
print("-" * 74)
print(f"  Total unique class names : {len(all_unique)}")
print(f"  Exact/user matches       : {len(exact_rows)}")
print(f"  Fuzzy matches            : {len(fuzzy_rows)}")
print(f"  New (fieldid=3)          : {sum(1 for _,_,f in new_rows if f==3)}")
print(f"  New (fieldid=2)          : {sum(1 for _,_,f in new_rows if f==2)}")
print("-" * 74)

# ── Competition blocks summary ────────────────────────────────────────────────
print("\n--- COMPETITION BLOCKS & DAY ASSIGNMENTS ---")
COLOR_NAMES = {
    "FFD9E1F2": "blue",
    "FFFCE4D6": "orange/peach",
    "FFFFF2CC": "yellow",
}
for b in comp_blocks:
    days   = b["days"]
    n_days = len(days)
    cg     = b["color_groups"]
    print(f"\n  [{b['sheet']}]  {b['comp_name']}")
    print(f"    dates : {b['start']} – {b['end']}  ({n_days} days)")
    print(f"    ranch : ranchid=11 (דאבל קיי)   fieldid={b['field_id']}")
    print(f"    color groups detected: {len(cg)}")
    for i, (color, day_idx) in enumerate(cg):
        cname = COLOR_NAMES.get(color, color)
        day_d = days[day_idx] if day_idx < len(days) else "?"
        print(f"      group {i+1}: {cname} ({color}) → day offset {day_idx} → {day_d}")
    if b["cutting_split"]:
        print(f"    Cutting block starts at row {b['cutting_split']} (separate competition)")

print()
print("─" * 74)
print("Phase 1 complete. Waiting for confirmation before Phase 2 insertion.")
print("─" * 74)
