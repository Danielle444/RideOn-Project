"""Phase 1 — Class Name Matching Report for historical data insertion."""
import sys
import subprocess

# Ensure UTF-8 output on Windows
sys.stdout.reconfigure(encoding="utf-8", errors="replace")

# ── Install dependencies if missing ─────────────────────────────────────────
for pkg in ("openpyxl", "python-dotenv", "supabase"):
    try:
        __import__(pkg.replace("-", "_").split("[")[0])
    except ImportError:
        print(f"Installing {pkg}...")
        subprocess.check_call([sys.executable, "-m", "pip", "install", pkg, "-q"])

import re
import difflib
from pathlib import Path
from dotenv import load_dotenv
import os
import openpyxl

# ── Credentials ──────────────────────────────────────────────────────────────
load_dotenv(Path(r"C:\Users\oren ahuvia\Documents\GitHub\RideOn-Project\.env"))
from supabase import create_client
sb = create_client(os.environ["SUPABASE_URL"], os.environ["SUPABASE_KEY"])

# ── Excel path ───────────────────────────────────────────────────────────────
EXCEL = Path(r"C:\Users\oren ahuvia\Documents\GitHub\RideOn-Project\סיכום תחרויות ריינינג 25.xlsx")

print(f"Loading: {EXCEL.name}")
wb = openpyxl.load_workbook(EXCEL, data_only=True)
sheet_names = wb.sheetnames
print(f"Sheets: {sheet_names}\n")

gilyonot  = [s for s in sheet_names if s.startswith("גיליון")]
summaries = [s for s in sheet_names if not s.startswith("גיליון")]
print(f"Class sheets (גיליון): {gilyonot}")
print(f"Summary sheets:        {summaries}\n")

# ── Skip rules ───────────────────────────────────────────────────────────────
SKIP_PREFIXES = ['פייד טיים', 'מספר', 'מקצה לרישום דמי ביטול', 'סה"כ']

def should_skip(name):
    if not name:
        return True
    n = str(name).strip()
    if not n:
        return True
    if n == "מספר":
        return True
    for p in SKIP_PREFIXES:
        if n.startswith(p):
            return True
    try:
        float(n)
        return True
    except ValueError:
        pass
    return False

# ── Collect all unique class names ───────────────────────────────────────────
all_class_names = set()
for sheet in gilyonot:
    ws   = wb[sheet]
    rows = list(ws.iter_rows(values_only=True))
    for row in rows[3:]:
        raw = row[0]
        if raw is None:
            continue
        name = str(raw).strip()
        if not should_skip(name):
            all_class_names.add(name)

print(f"Unique class names found: {len(all_class_names)}\n")

# ── Fetch existing classtypes ────────────────────────────────────────────────
print("Fetching classtypes from Supabase...")
existing = sb.table("classtype").select("classtypeid, classname, fieldid").execute().data
print(f"Found {len(existing)} existing classtypes\n")

db_names = [r["classname"] for r in existing]
db_map   = {r["classname"]: r for r in existing}

# ── Match ────────────────────────────────────────────────────────────────────
exact_matches  = []
fuzzy_matches  = []
no_matches     = []

for name in sorted(all_class_names):
    # exact (case-insensitive)
    hit = next((db_map[d] for d in db_names if d.strip().lower() == name.strip().lower()), None)
    if hit:
        exact_matches.append((name, hit))
        continue
    # fuzzy
    close = difflib.get_close_matches(name, db_names, n=1, cutoff=0.6)
    if close:
        score = round(difflib.SequenceMatcher(None, name, close[0]).ratio(), 2)
        fuzzy_matches.append((name, db_map[close[0]], close[0], score))
    else:
        no_matches.append(name)

# ── Report ───────────────────────────────────────────────────────────────────
print("=" * 62)
print("CLASS NAME MATCHING REPORT")
print("=" * 62)

for name, ct in exact_matches:
    print(f"  [EXACT] {name:<42} -> id={ct['classtypeid']}")

print()
for name, ct, db_name, score in fuzzy_matches:
    print(f"  [FUZZY] {name:<42} -> id={ct['classtypeid']} ({db_name}) [score={score}]")

print()
for name in no_matches:
    print(f"  [NEW]   {name}")

print()
print("-" * 62)
print(f"  Total unique:   {len(all_class_names)}")
print(f"  Exact matches:  {len(exact_matches)}")
print(f"  Fuzzy matches:  {len(fuzzy_matches)}")
print(f"  No match (NEW): {len(no_matches)}")
print("-" * 62)

if no_matches or fuzzy_matches:
    print("\nClasses needing review:")
    for n in no_matches:
        print(f"  [NEW]   {n}")
    for name, ct, db_name, score in fuzzy_matches:
        print(f"  [FUZZY] {name}  ->  {db_name} [score={score}]")

print()
print("─" * 62)
print("Phase 1 complete. Review and confirm to proceed with Phase 2.")
print("─" * 62)
