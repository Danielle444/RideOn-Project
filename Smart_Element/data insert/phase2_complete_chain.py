"""
Phase 2 Chain Completer
For orphaned persons (created but chain never finished), completes
person → federationmember → horse → systemuser → bill → servicerequest → entry.
Falls back to fresh creation if person doesn't exist yet.
"""
import sys
import subprocess

sys.stdout.reconfigure(encoding="utf-8", errors="replace")

for pkg in ("openpyxl", "python-dotenv", "supabase"):
    try:
        __import__(pkg.replace("-", "_").split("[")[0])
    except ImportError:
        subprocess.check_call([sys.executable, "-m", "pip", "install", pkg, "-q"])

import re
from pathlib import Path
from datetime import datetime, date, time, timedelta
from zoneinfo import ZoneInfo
from dotenv import load_dotenv
import os
import openpyxl

load_dotenv(Path(r"C:\Users\oren ahuvia\Documents\GitHub\RideOn-Project\.env"))
from supabase import create_client
sb = create_client(os.environ["SUPABASE_URL"], os.environ["SUPABASE_KEY"])

HOST_RANCH_ID = 11
TZ            = ZoneInfo("Asia/Jerusalem")
EXCEL         = Path(r"C:\Users\oren ahuvia\Documents\GitHub\RideOn-Project\סיכום תחרויות ריינינג 25.xlsx")

CLASSTYPE_MAP = {
    "ירוקי התאחדות": 2, "ירוקי רוכב חדש התאחדות": 3, "נוביס התאחדות": 10,
    "נוביס נונ פרו התאחדות": 11, "נונ פרו 50+": 6, "נוער ירוקי התאחדות": 4,
    "פתוח לא מוגבל": 1, "prime time non-professional riders": 14,
    "נוער 14-18 nrha": 15, "פתוח לא מוגבל b2w": 1, "פתוח לא מוגבל dk": 1,
    "irbc דרבי נונ פרו l1": 123, "irbc דרבי נונ פרו l3": 124,
    "irbc דרבי נונ פרו l4": 125, "irbc דרבי נונ פרו פריים טיים": 126,
    "irbc דרבי פתוח l1": 127, "irbc דרבי פתוח l3": 128, "irbc דרבי פתוח l4": 129,
    "junior riders": 130, "senior non-professional riders": 131,
    "senior professional riders": 132, "young non-professional riders": 133,
    "young professional riders": 134, "אקסטרים שוטאווט": 135,
    "דרבי נונ פרו l4": 136, "דרבי פתוח l4": 137, "זוגות 1": 138, "זוגות 2": 139,
    "נוביס l1 nrha": 140, "נונ פרו nrha": 141, "נונ פרו מוגבל nrha": 142,
    "נונ פרו פריים טיים 40+ nrha": 143, "נוער 13 nrha": 144,
    "נוער unrestricted nrha": 145, "פטוריטי נונ פרו nrha": 146,
    "פטוריטי נונ פרו בני 4 irbc l1": 147, "פטוריטי נונ פרו בני 4 irbc l2": 148,
    "פטוריטי נונ פרו בני 4 irbc l3": 149, "פטוריטי נונ פרו בני 4 irbc l4": 150,
    "פטוריטי פתוח nrha": 151, "פטוריטי פתוח בני 3 irbc l1": 152,
    "פטוריטי פתוח בני 3 irbc l2": 153, "פטוריטי פתוח בני 3 irbc l3": 154,
    "פטוריטי פתוח בני 3 irbc l4": 155, "פתוח nrha": 156, "פתוח מוגבל nrha": 157,
}

def resolve(name):
    return CLASSTYPE_MAP.get(str(name).strip().lower())

SKIP_PREFIXES = ["פייד טיים", "מספר", "מקצה לרישום דמי ביטול", 'סה"כ']
def should_skip(name):
    if not name: return True
    n = str(name).strip()
    if not n or n == "מספר": return True
    for p in SKIP_PREFIXES:
        if n.startswith(p): return True
    try: float(n); return True
    except ValueError: pass
    return False

def is_paid_time(name):
    return bool(name) and str(name).strip().startswith("פייד טיים")

DAY_OFFSETS = {"שני": 1, "שלישי": 2, "רביעי": 3, "חמישי": 4, "שישי": 5}
def day_offset(text):
    for d, o in DAY_OFFSETS.items():
        if d in str(text): return o
    return 0

def parse_dates(cell):
    s = str(cell).strip()
    m = re.search(r"(\d+)-(\d+)\.(\d+)\.(\d+)", s)
    if m: return date(int(m[4]), int(m[3]), int(m[1])), date(int(m[4]), int(m[3]), int(m[2]))
    m2 = re.search(r"(\d+)\.(\d+)\.(\d+)", s)
    if m2: d_ = date(int(m2[3]), int(m2[2]), int(m2[1])); return d_, d_
    raise ValueError(f"Cannot parse: {cell!r}")

def to_float(v):
    try: return float(v) if v not in (None, "") else 0.0
    except: return 0.0

# ── Load all existing historical nationalids ─────────────────────────────────
print("Loading existing historical nationalids...")
hist_rows = sb.table("person").select("nationalid, personid").like("nationalid", "000%").execute().data
used_map  = {int(r["nationalid"]): r["personid"] for r in hist_rows if r["nationalid"].isdigit()}
next_fresh = max(used_map.keys()) + 1 if used_map else 10000
print(f"  {len(used_map)} historical persons exist, next fresh id: {next_fresh:09d}\n")

# ── Core: complete one entry, reusing person if it already exists ─────────────
def make_entry(n, cic_id, cls_dt, comp_id):
    nat = f"{n:09d}"

    # person
    if n in used_map:
        pid = used_map[n]
    else:
        p   = sb.table("person").insert({
            "nationalid": nat, "firstname": "רוכב", "lastname": f"היסטורי {n}",
        }).execute().data[0]
        pid = p["personid"]
        used_map[n] = pid

    # federationmember (idempotent)
    if not sb.table("federationmember").select("federationmemberid").eq("federationmemberid", pid).execute().data:
        sb.table("federationmember").insert({"federationmemberid": pid, "hasvalidmembership": True}).execute()

    # horse (always new — no unique constraint)
    h = sb.table("horse").insert({"ranchid": HOST_RANCH_ID, "horsename": f"סוס היסטורי {n}"}).execute().data[0]

    # systemuser (idempotent)
    if not sb.table("systemuser").select("systemuserid").eq("systemuserid", pid).execute().data:
        sb.table("systemuser").insert({
            "systemuserid": pid, "username": f"hist_{n}",
            "passwordhash": "placeholder", "passwordsalt": "placeholder", "isactive": False,
        }).execute()

    # bill
    bill = sb.table("bill").insert({
        "paidbypersonid": pid, "amounttopay": 0,
        "dateopened": cls_dt.isoformat(), "competitionid": comp_id,
    }).execute().data[0]

    # servicerequest
    sr = sb.table("servicerequest").insert({
        "orderedbysystemuserid": pid, "horseid": h["horseid"],
        "riderfederationmemberid": pid, "billid": bill["billid"],
    }).execute().data[0]

    # entry
    sb.table("entry").insert({
        "entryid": sr["srequestid"], "classincompid": cic_id, "entrystatus": "Active",
    }).execute()

# ── Parse Excel to find shortfalls ───────────────────────────────────────────
print("Scanning Excel for shortfalls...")
wb       = openpyxl.load_workbook(EXCEL, data_only=True)
ns_      = wb.sheetnames
gilyonot = [s for s in ns_ if s.startswith("גיליון")]

needed   = []   # (cic_id, shortfall, cls_dt, comp_id, counter_start)
counter  = 10000  # replays same counter logic to find the right n-values

for sheet in gilyonot:
    ws    = wb[sheet]
    rows_ = list(ws.iter_rows(values_only=True))

    comp_name              = str(rows_[0][0]).strip()
    event_start, event_end = parse_dates(rows_[1][0])

    comp_r = sb.table("competition").select("competitionid").eq("competitionname", comp_name).execute()
    if not comp_r.data: continue
    comp_id = comp_r.data[0]["competitionid"]

    hdr = rows_[2]
    ci_ent = ci_org = ci_fed = None
    for i, h in enumerate(hdr):
        if h is None: continue
        hs = str(h).strip()
        if "מספר כניסות" in hs: ci_ent = i
        elif hs == "חווה":      ci_org = i
        elif hs == "התאחדות":   ci_fed = i

    cur_off = 0
    for row in rows_[3:]:
        if row[0] is None: continue
        raw = str(row[0]).strip()
        if is_paid_time(raw):
            cur_off = day_offset(raw); continue
        if should_skip(raw): continue

        ctype_id = resolve(raw)
        if ctype_id is None: continue

        n_entries = int(to_float(row[ci_ent])) if ci_ent is not None else 0
        cls_date  = event_start + timedelta(days=cur_off)
        cls_dt    = datetime.combine(cls_date, time(9, 0)).replace(tzinfo=TZ)
        date_iso  = cls_date.isoformat()

        cic_r = (
            sb.table("classincompetition").select("classincompid")
            .eq("competitionid", comp_id).eq("classtypeid", ctype_id)
            .gte("classdatetime", f"{date_iso}T00:00:00+00:00")
            .lte("classdatetime", f"{date_iso}T23:59:59+00:00")
            .execute()
        )
        if not cic_r.data:
            if n_entries > 0: counter += n_entries
            continue
        cic_id = cic_r.data[0]["classincompid"]

        if n_entries > 0:
            ex_e  = sb.table("entry").select("entryid").eq("classincompid", cic_id).execute()
            have  = len(ex_e.data)
            if have < n_entries:
                shortfall = n_entries - have
                start_n   = counter + have   # where this class's missing entries should start
                needed.append((cic_id, shortfall, cls_dt, comp_id, start_n))
                print(f"  cic={cic_id}: have={have}/{n_entries}, need {shortfall} more (n starts at {start_n:09d})")
            counter += n_entries
        # n_entries==0: no counter advance

print(f"\n  {len(needed)} classes need entries\n")

if not needed:
    print("All classes are complete.")
    import sys; sys.exit(0)

# ── Fabricate ────────────────────────────────────────────────────────────────
print("Completing entry chains...")
total = 0; errors = 0

for cic_id, shortfall, cls_dt, comp_id, start_n in needed:
    for i in range(shortfall):
        n = start_n + i
        # If this n is already used but has no entry for this cic, we reuse.
        # make_entry handles idempotent person/fm/su creation.
        try:
            make_entry(n, cic_id, cls_dt, comp_id)
            total += 1
        except Exception as e:
            # If even reuse fails, try a fresh id beyond the known range
            try:
                make_entry(next_fresh, cic_id, cls_dt, comp_id)
                print(f"  Fallback: used fresh id {next_fresh:09d} for cic={cic_id}")
                next_fresh += 1
                total += 1
            except Exception as e2:
                print(f"  ERROR cic={cic_id} n={n}: {e} / {e2}")
                errors += 1

        if total > 0 and total % 50 == 0:
            print(f"  ... {total} entries fabricated")

    print(f"  cic={cic_id}: {shortfall} entries completed")

print()
print("=" * 60)
print("CHAIN COMPLETION DONE")
print("=" * 60)
print(f"  Entries fabricated : {total}")
print(f"  Errors             : {errors}")
print("=" * 60)
