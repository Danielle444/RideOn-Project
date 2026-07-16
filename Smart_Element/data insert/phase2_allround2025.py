"""
Phase 2 — Full insertion for All-Around 2025 (Format D).
Applies all user-confirmed corrections from Phase 1.
"""
import sys
import subprocess

sys.stdout.reconfigure(encoding="utf-8", errors="replace")

for pkg in ("openpyxl", "pandas", "python-dotenv", "supabase"):
    try:
        __import__(pkg.replace("-", "_").split("[")[0])
    except ImportError:
        subprocess.check_call([sys.executable, "-m", "pip", "install", pkg, "-q"])

import re
from pathlib import Path
from datetime import datetime, date, time, timedelta
from zoneinfo import ZoneInfo
from dotenv import load_dotenv
import os
import pandas as pd
from openpyxl import load_workbook

load_dotenv(Path(r"C:\Users\oren ahuvia\Documents\GitHub\RideOn-Project\.env"))
from supabase import create_client
sb = create_client(os.environ["SUPABASE_URL"], os.environ["SUPABASE_KEY"])

EXCEL       = Path(r"C:\Users\oren ahuvia\Documents\GitHub\RideOn-Project\תחרויות אולארונד 2025.xlsx")
TZ          = ZoneInfo("Asia/Jerusalem")
RANCH_ID    = 11
ARENA_ID    = 2
CREATOR_UID = 2
COUNTER_START = 40000

# ── Skip rule: words that mark skippable rows ─────────────────────────────────
SKIP_WORDS  = ["סירקט", "שואוטאוט", "שוטאאוט", "שוטאווט", "תמריצים", "זוגות"]
SKIP_STARTS = ["מקצה לרישום דמי ביטול", "מקצה לביטול", 'סה"כ', "סה\"כ"]

def should_skip(raw):
    if not raw or str(raw).strip() in ("", "nan"):
        return True
    n = str(raw).strip()
    if not n or n == "מספר":
        return True
    for s in SKIP_STARTS:
        if n.startswith(s):
            return True
    for w in SKIP_WORDS:
        if w in n:
            return True
    try:
        float(n)
        return True
    except ValueError:
        pass
    return False

# ── Typo corrections (apply BEFORE lookup) ────────────────────────────────────
TYPO_MAP = {
    "הורסמני פתוח":                    "הורסמנשיפ פתוח",
    "הורסמשניפ שוטאאוט 20":            "הורסמנשיפ שוטאאוט",
    "טרייל שוטאאוט 20":                "טרייל שוטאאוט",
    "האנט סיט אקויטיישן עד גיל 13 ירוקי": "האנט סיט אקוויטיישן עד גיל 13 ירוקי",
    "האנט סיט אקויטיישן":               "האנט סיט אקוויטיישן",
}

def normalize(raw):
    if not isinstance(raw, str):
        return ""
    n = raw.strip()
    n = TYPO_MAP.get(n, n)
    return re.sub(r"\s+", " ", n).strip()

# ── All-Around class name → classtypeid (fieldid=3) ───────────────────────────
ALLROUND_MAP = {
    # Exact matches from Phase 1
    "האנט סיט אקוויטיישן ירוקי עד 15":  78,
    "האנט סיט אקוויטיישן עד גיל 13":    80,
    "האנט סיט אקוויטיישן עד גיל 18":    74,
    "האנט סיט אקוויטיישן פתוח":         72,
    "האנטר אנדר סאדל 13 ירוקי":         82,
    "האנטר אנדר סאדל ירוקי עד 15":      79,
    "האנטר אנדר סאדל ירוקי עד 18":      76,
    "האנטר אנדר סאדל עד גיל 13":        81,
    "האנטר אנדר סאדל עד גיל 15":        77,
    "האנטר אנדר סאדל עד גיל 18":        75,
    "האנטר אנדר סאדל פתוח":             73,
    "הורסמנשיפ הליכה ג'וג עד גיל 18":   70,
    "הורסמנשיפ ירוקי בוגרים":           89,
    "הורסמנשיפ ירוקי עד 13":            90,
    "הורסמנשיפ ירוקי עד 15":            114,
    "הורסמנשיפ ירוקי עד 18":            71,
    "הורסמנשיפ נוביס נוביס":            92,
    "הורסמנשיפ נונ פרו":                 110,
    "הורסמנשיפ נונ פרו 40+ הליכה ג'וג": 112,
    "הורסמנשיפ עד 15":                  91,
    "הורסמנשיפ עד 18":                  113,
    "הורסמנשיפ עד גיל 10":              93,
    "הורסמנשיפ עד גיל 13":              115,
    "הורסמנשיפ פתוח":                   88,
    "הורסמנשיפ פתוח לסוסי נוביס":       111,
    "הליכה ג'וג סירקט עד 13":           101,  # exists but will be skipped (contains סירקט)
    "הליכה ג'וג עד גיל 10":             96,
    "הליכה ג'וג עד גיל 13":             120,
    "הליכה ג'וג עד גיל 18":             69,
    "טרייל הליכה ג'וג עד גיל 18":       86,
    "טרייל ירוקי בוגרים":               107,
    "טרייל ירוקי עד 13":                108,
    "טרייל ירוקי עד 15":                66,
    "טרייל ירוקי עד 18":                85,
    "טרייל נוביס נוביס":                105,
    "טרייל נונ פרו":                     83,
    "טרייל עד 15":                      106,
    "טרייל עד 18":                      65,
    "טרייל עד גיל 10":                  109,
    "טרייל עד גיל 13":                  87,
    "טרייל פתוח":                       64,
    "טרייל פתוח לסוסי נוביס":           84,
    "מקצה אימון הורסמנשיפ":             67,
    "מקצה אימון טרייל":                 63,
    "מקצה אימון פלז'ר":                 68,
    "פלז'ר ירוקי בוגרים":               97,
    "פלז'ר ירוקי עד 15":                100,
    "פלז'ר ירוקי עד 18":                118,
    "פלז'ר נונ פרו":                     94,
    "פלז'ר נונ פרו 40+ הליכה ג'וג":     95,
    "פלז'ר עד 15":                      119,
    "פלז'ר עד 18":                      99,
    "פלז'ר עד גיל 13":                  102,
    "פלז'ר פתוח":                       116,
    "פלז'ר פתוח לסוסי נוביס":           98,
    "פלזר נוביס נוביס":                  117,
    "שואומנשיפ":                        103,
    "שואומנשיפ ירוקי":                  104,
    # User corrections
    "הורסמנשיפ 13":                      115,
    # נוביס → עד 18 equivalents (user rule)
    "האנטר אנדר סאדל נוביס":            75,    # → עד גיל 18
    "הורסמנשיפ נוביס":                   111,   # user override
    "טרייל נוביס":                       65,    # → עד 18
    "טרייל ירוקי נוביס":                 85,    # → ירוקי עד 18
    "פלז'ר ירוקי נוביס":                 118,   # → ירוקי עד 18
    "פלז'ר נוביס":                       98,    # user override
    "פלזר נוביס":                        98,    # user override (variant spelling)
    # NEW classtypes (IDs filled after Step 0):
    # "האנט סיט אקוויטיישן עד גיל 13 ירוקי" → filled below
    # "הורסמנשיפ נוביס ירוקי" → filled below
    # "ווסטרן ריידינג" / "ראנץ ריידינג" → filled below
    # "טרייל פתוח בלי מתג" → filled below
    # "הורסמנשיפ בלי אוכף" → filled below
}

# ── Cutting class name → classtypeid (fieldid=2) ─────────────────────────────
CUTTING_MAP = {
    "קאוהורס נונ פרו":  17,
    "נוביס":            22,
    "נוביס נונ פרו":    23,
    "פתוח":             18,
    "פתוח ncha":        19,
    "נונ פרו":           20,
    "נונ פרו ncha":     21,
    "קאוהורס פתוח":     26,
    "נונ פרו פלאטינום":  27,
    "מוגבל ncha":       29,
    "נוביס ncha":       22,   # missed in phase 1 → map to קאטינג נוביס
    "עד 18 ירוקי":      32,
    "עד 15 ירוקי":      33,
    "קאטינג ירוקי":     34,
}

def resolve_class(norm, field_id):
    """Return classtypeid or None (=skip)."""
    key = norm.lower()
    if field_id == 3:
        return ALLROUND_MAP.get(norm)  # exact key match (already normalized)
    elif field_id == 2:
        return CUTTING_MAP.get(key)
    return None

# ── Standard jackpot prizes (prizetypeid=2, amount=100) ──────────────────────
JACKPOT_CLASSES = {
    64,   # טרייל פתוח
    83,   # טרייל נונ פרו
    84,   # טרייל פתוח לסוסי נוביס
    107,  # טרייל ירוקי בוגרים
    72,   # האנט סיט אקוויטיישן פתוח
    73,   # האנטר אנדר סאדל פתוח
    88,   # הורסמנשיפ פתוח
    89,   # הורסמנשיפ ירוקי בוגרים
    110,  # הורסמנשיפ נונ פרו
    111,  # הורסמנשיפ פתוח לסוסי נוביס
    116,  # פלז'ר פתוח
    94,   # פלז'ר נונ פרו
    97,   # פלז'ר ירוקי בוגרים
    98,   # פלז'ר פתוח לסוסי נוביס
}

# ── Winter-specific prizes: (classtypeid, day_offset) → [(prizetypeid, amount)]
# Note: some Winter prizes reference classes by their mapped classtypeid+day
WINTER_PRIZES_BY_TYPE_AND_DAY = {
    # Day 0 (Dec 11, yellow):
    (88, 0):  [(3, 1000)],   # הורסמנשיפ פתוח ₪1,000
    (75, 0):  [(3, 1000)],   # האנטר אנדר סאדל נוביס→עד גיל 18 ₪1,000
    # Day 1 (Dec 12, blue):
    (98, 1):  [(3, 700)],    # פלז'ר נוביס → פלז'ר פתוח לסוסי נוביס ₪700
    # Day 2 (Dec 13, orange):
    (65, 2):  [(3, 700)],    # טרייל נוביס → טרייל עד 18 ₪700
    (111, 2): [(3, 700)],    # הורסמנשיפ נוביס ₪700
    (116, 2): [(3, 1000)],   # פלז'ר פתוח ₪1,000
    # NEW classtype prizes (filled after creation):
    # טרייל פתוח בלי מתג day 2 → ₪1,500
    # הורסמנשיפ בלי אוכף day 2 → ₪1,500
}

def to_float(v):
    try:
        return float(str(v).strip()) if v not in (None, "") and str(v).strip() not in ("nan",) else 0.0
    except Exception:
        return 0.0

def get_row_color(ws, row_idx):
    for cell in ws[row_idx]:
        if cell.fill and cell.fill.fgColor:
            fc = cell.fill.fgColor
            if fc.type == "rgb" and fc.rgb not in ("00000000", "FF000000"):
                return fc.rgb
    return None

def parse_date_range(s):
    m = re.search(r"(\d{1,2})[-–](\d{1,2})[./](\d{1,2})[./](\d{2,4})", str(s))
    if not m:
        return None, None
    d1, d2, mo, yr = int(m[1]), int(m[2]), int(m[3]), int(m[4])
    if yr < 100:
        yr += 2000
    return date(yr, mo, d1), date(yr, mo, d2)

CUTTING_SPLIT_RE = re.compile(r"קאטינג\s+\d+\s+\d{1,2}[-–]\d{1,2}\.\d{2}")

# ════════════════════════════════════════════════════════════════════════════
# STEP 0  Cleanup + Create new classtypes
# ════════════════════════════════════════════════════════════════════════════
print("STEP 0 — Cleanup and new classtypes")
print("─" * 62)

# Delete any classincompetition records for classtypes containing זוגות
zoget_types = sb.table("classtype").select("classtypeid, classname").ilike("classname", "%זוגות%").execute().data
for zt in zoget_types:
    cics = sb.table("classincompetition").select("classincompid").eq("classtypeid", zt["classtypeid"]).execute().data
    for cic in cics:
        cic_id = cic["classincompid"]
        # Delete entries first
        entries = sb.table("entry").select("entryid").eq("classincompid", cic_id).execute().data
        for e in entries:
            try:
                sr = sb.table("servicerequest").select("srequestid, billid, horseid, riderfederationmemberid").eq("srequestid", e["entryid"]).execute().data
                sb.table("entry").delete().eq("entryid", e["entryid"]).execute()
                if sr:
                    sr = sr[0]
                    bill_r = sb.table("bill").select("paidbypersonid").eq("billid", sr["billid"]).execute().data
                    pid = bill_r[0]["paidbypersonid"] if bill_r else None
                    sb.table("servicerequest").delete().eq("srequestid", sr["srequestid"]).execute()
                    sb.table("bill").delete().eq("billid", sr["billid"]).execute()
                    if pid:
                        sb.table("systemuser").delete().eq("systemuserid", pid).execute()
                        sb.table("horse").delete().eq("horseid", sr["horseid"]).execute()
                        sb.table("federationmember").delete().eq("federationmemberid", pid).execute()
                        sb.table("person").delete().eq("personid", pid).execute()
            except Exception as ex:
                print(f"  Error deleting entry chain for cic={cic_id}: {ex}")
        sb.table("classprize").delete().eq("classincompid", cic_id).execute()
        sb.table("classincompetition").delete().eq("classincompid", cic_id).execute()
        print(f"  DELETED classincompid={cic_id} (classtype '{zt['classname']}')")

# New classtypes to create
NEW_TYPES = [
    (3, "האנט סיט אקוויטיישן עד גיל 13 ירוקי"),
    (3, "הורסמנשיפ נוביס ירוקי"),
    (3, "ווסטרן ריידינג"),
    (3, "טרייל פתוח בלי מתג"),
    (3, "הורסמנשיפ בלי אוכף"),
]

existing_all = sb.table("classtype").select("classtypeid, classname").execute().data
existing_lower = {r["classname"].strip().lower(): r["classtypeid"] for r in existing_all}

new_types_created = 0
for fid, cname in NEW_TYPES:
    key = cname.strip().lower()
    if key in existing_lower:
        cid = existing_lower[key]
        print(f"  ALREADY EXISTS: '{cname}' → id={cid}")
    else:
        try:
            r = sb.table("classtype").insert({"fieldid": fid, "classname": cname}).execute()
            cid = r.data[0]["classtypeid"]
            print(f"  NEW classtype: '{cname}' → id={cid} (fieldid={fid})")
            new_types_created += 1
        except Exception as e:
            print(f"  ERROR creating '{cname}': {e}")
            cid = None

    # Wire into maps
    if cid:
        if cname == "האנט סיט אקוויטיישן עד גיל 13 ירוקי":
            ALLROUND_MAP[cname] = cid
        elif cname == "הורסמנשיפ נוביס ירוקי":
            ALLROUND_MAP[cname] = cid
        elif cname == "ווסטרן ריידינג":
            ALLROUND_MAP["ווסטרן ריידינג"] = cid
            ALLROUND_MAP["ראנץ ריידינג"] = cid
            ALLROUND_MAP["ראנץ ריידינג פתוח"] = cid
        elif cname == "טרייל פתוח בלי מתג":
            ALLROUND_MAP[cname] = cid
            WINTER_PRIZES_BY_TYPE_AND_DAY[(cid, 2)] = [(3, 1500)]
        elif cname == "הורסמנשיפ בלי אוכף":
            ALLROUND_MAP[cname] = cid
            WINTER_PRIZES_BY_TYPE_AND_DAY[(cid, 2)] = [(3, 1500)]

print(f"\n  {new_types_created} new classtypes created\n")

# ════════════════════════════════════════════════════════════════════════════
# Load Excel
# ════════════════════════════════════════════════════════════════════════════
print("Loading Excel...")
wb      = load_workbook(EXCEL, data_only=True)
all_dfs = pd.read_excel(EXCEL, sheet_name=None, header=None, dtype=str)
print(f"  Sheets: {list(wb.sheetnames)}\n")

entries_todo = []   # (classincompid, n_entries, cls_dt, comp_id, ranchid)
S = dict(comps_ins=0, comps_skip=0, cls_ins=0, cls_skip=0,
         prizes=0, entries=0, rows_skip=0, new_types=new_types_created, errors=0)

def col_of(df, hdr_row, *keywords):
    if hdr_row >= len(df): return None
    for i, v in enumerate(df.iloc[hdr_row]):
        if pd.isna(v): continue
        for kw in keywords:
            if kw in str(v).strip(): return i
    return None

def process_allround_block(sheet_name, ws, df, comp_name, start_d, end_d,
                            field_id, is_winter, hdr_row, data_start, data_end,
                            prize_source_key):
    """Process one competition block and insert classincompetition + prizes."""
    is_cutting = (field_id == 2)

    # Build color → day map (dynamic: order of first appearance)
    color_to_day = {}
    day_seq = [0]
    day_dates = []
    cur = start_d
    while cur <= end_d:
        day_dates.append(cur)
        from datetime import timedelta as td
        cur += td(days=1)

    for row in ws.iter_rows(min_row=data_start + 1, max_row=data_end):
        color = get_row_color(ws, row[0].row)
        if color and color not in color_to_day:
            idx = min(day_seq[0], len(day_dates) - 1)
            color_to_day[color] = idx
            day_seq[0] += 1

    # Competition insert
    ex = sb.table("competition").select("competitionid").eq("competitionname", comp_name).execute()
    if ex.data:
        comp_id = ex.data[0]["competitionid"]
        print(f"    SKIP competition (exists id={comp_id})")
        S["comps_skip"] += 1
    else:
        try:
            r = sb.table("competition").insert({
                "hostranchid":           RANCH_ID,
                "fieldid":               field_id,
                "createdbysystemuserid": CREATOR_UID,
                "competitionname":       comp_name,
                "competitionstartdate":  start_d.isoformat(),
                "competitionenddate":    end_d.isoformat(),
                "competitionstatus":     "הסתיימה",
            }).execute()
            comp_id = r.data[0]["competitionid"]
            print(f"    INSERTED competition id={comp_id}")
            S["comps_ins"] += 1
        except Exception as e:
            print(f"    ERROR competition: {e}")
            S["errors"] += 1
            return

    ci_ent = col_of(df, hdr_row, "מספר כניסות")
    ci_org = col_of(df, hdr_row, "חווה")
    ci_fed = col_of(df, hdr_row, "התאחדות")
    ci_prz = None
    if is_cutting:
        hdr = df.iloc[hdr_row]
        last = max((i for i in range(len(hdr)) if pd.notna(hdr.iloc[i]) and str(hdr.iloc[i]).strip()), default=None)
        if last and last + 1 < len(df.columns):
            ci_prz = last + 1

    day_order_counter = {}
    pending_prizes = {}  # classincompid → [(prizetypeid, amount)]

    for ri in range(data_start, data_end):
        if ri >= len(df):
            break
        raw = df.iloc[ri, 0]
        if pd.isna(raw):
            continue
        raw_s = str(raw).strip()
        if should_skip(raw_s):
            S["rows_skip"] += 1
            continue

        norm = normalize(raw_s)
        if not norm:
            continue

        ctype_id = resolve_class(norm, field_id)
        if ctype_id is None:
            S["rows_skip"] += 1
            continue

        # Day from row color (openpyxl row = ri + 2 when data_start=2 in pandas)
        openpyxl_row = ri + 1   # pandas 0-indexed → openpyxl 1-indexed
        color = get_row_color(ws, openpyxl_row)
        day_idx = color_to_day.get(color, 0) if color else 0
        cls_date = day_dates[day_idx] if day_idx < len(day_dates) else day_dates[-1]
        cls_dt   = datetime.combine(cls_date, time(9, 0)).replace(tzinfo=TZ)
        date_iso = cls_date.isoformat()

        day_order_counter[day_idx] = day_order_counter.get(day_idx, 0) + 1
        order = day_order_counter[day_idx]

        n_entries = int(to_float(df.iloc[ri, ci_ent])) if ci_ent is not None else 0
        org_cost  = to_float(df.iloc[ri, ci_org])      if ci_org is not None else 0.0
        fed_cost  = to_float(df.iloc[ri, ci_fed])      if ci_fed is not None else 0.0

        # Idempotency
        ex_cls = (
            sb.table("classincompetition").select("classincompid")
            .eq("competitionid", comp_id).eq("classtypeid", ctype_id)
            .gte("classdatetime", f"{date_iso}T00:00:00+00:00")
            .lte("classdatetime", f"{date_iso}T23:59:59+00:00")
            .execute()
        )
        if ex_cls.data:
            cic_id = ex_cls.data[0]["classincompid"]
            S["cls_skip"] += 1
            if n_entries > 0:
                ex_e = sb.table("entry").select("entryid").eq("classincompid", cic_id).execute()
                if len(ex_e.data) < n_entries:
                    entries_todo.append((cic_id, n_entries, cls_dt, comp_id, RANCH_ID))
        else:
            try:
                r = sb.table("classincompetition").insert({
                    "competitionid":  comp_id,
                    "classtypeid":    ctype_id,
                    "arenaranchid":   RANCH_ID,
                    "arenaid":        ARENA_ID,
                    "classdatetime":  cls_dt.isoformat(),
                    "organizercost":  org_cost,
                    "federationcost": fed_cost,
                    "orderinday":     order,
                }).execute()
                cic_id = r.data[0]["classincompid"]
                S["cls_ins"] += 1
                if n_entries > 0:
                    entries_todo.append((cic_id, n_entries, cls_dt, comp_id, RANCH_ID))
            except Exception as e:
                print(f"    ERROR class '{norm}': {e}")
                S["errors"] += 1
                continue

        # Prize determination
        prize_list = []
        if is_winter:
            winter_key = (ctype_id, day_idx)
            prize_list = WINTER_PRIZES_BY_TYPE_AND_DAY.get(winter_key, [])
        elif not is_cutting and ctype_id in JACKPOT_CLASSES:
            prize_list = [(2, 100)]
        elif is_cutting and ci_prz is not None and ci_prz < len(df.columns):
            ann = df.iloc[ri, ci_prz] if ci_prz < len(df.iloc[ri]) else None
            if not pd.isna(ann) if ann is not None else False:
                text = str(ann).strip()
                if text and text != "nan":
                    m = re.search(r"(\d+)", text)
                    amount = int(m.group(1)) if m else 0
                    if amount > 0:
                        if "ג'קפוט" in text:
                            prize_list.append((2, amount))
                        elif "תלושים" in text or "שובר" in text:
                            prize_list.append((1, amount))
                        elif "כסף מוסף" in text:
                            prize_list.append((3, amount))

        if prize_list:
            pending_prizes[cic_id] = prize_list

    # Insert prizes
    for cic_id, plist in pending_prizes.items():
        for ptype, amount in plist:
            try:
                ex = sb.table("classprize").select("classincompid").eq("classincompid", cic_id).eq("prizetypeid", ptype).execute()
                if not ex.data:
                    sb.table("classprize").insert({"classincompid": cic_id, "prizetypeid": ptype, "prizeamount": amount}).execute()
                    S["prizes"] += 1
            except Exception as e:
                print(f"    ERROR prize cic={cic_id} type={ptype}: {e}")
                S["errors"] += 1


# ════════════════════════════════════════════════════════════════════════════
# STEP 1+2  Process all sheets
# ════════════════════════════════════════════════════════════════════════════
print("STEP 1+2 — Competitions and Classes")
print("─" * 62)

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    df = all_dfs[sheet_name]

    comp_name  = str(df.iloc[0, 0]).strip() if pd.notna(df.iloc[0, 0]) else sheet_name
    date_raw   = str(df.iloc[1, 0]).strip() if len(df) > 1 and pd.notna(df.iloc[1, 0]) else ""
    start_d, end_d = parse_date_range(date_raw)
    if not start_d:
        print(f"\n  [{sheet_name}] — no date found, skipping")
        continue

    is_winter = "חורף" in sheet_name
    print(f"\n  [{sheet_name}]  {comp_name}  ({start_d}–{end_d})  winter={is_winter}")

    # Find Cutting block mid-sheet
    cutting_start_row = None
    for ri in range(3, len(df)):
        raw_val = str(df.iloc[ri, 0]).strip() if pd.notna(df.iloc[ri, 0]) else ""
        if CUTTING_SPLIT_RE.match(raw_val):
            cutting_start_row = ri
            break

    # Process All-Around block (rows 2 header, data rows 3 to cutting_start or end)
    data_end_ar = cutting_start_row if cutting_start_row else len(df)
    process_allround_block(
        sheet_name, ws, df, comp_name, start_d, end_d,
        field_id=3, is_winter=is_winter,
        hdr_row=2, data_start=3, data_end=data_end_ar,
        prize_source_key=sheet_name,
    )

    # Process Cutting block if present
    if cutting_start_row:
        cut_date_str = str(df.iloc[cutting_start_row, 0]).strip()
        cut_start, cut_end = parse_date_range(cut_date_str)
        m_num = re.search(r"קאטינג\s+(\d+)", cut_date_str)
        cut_name = f"תחרות קאטינג {m_num.group(1)} 2025" if m_num else f"תחרות קאטינג מ{cut_date_str}"

        if cut_start:
            print(f"\n    [Cutting block]  {cut_name}  ({cut_start}–{cut_end})")
            process_allround_block(
                sheet_name, ws, df, cut_name, cut_start, cut_end,
                field_id=2, is_winter=False,
                hdr_row=cutting_start_row + 1,
                data_start=cutting_start_row + 2,
                data_end=len(df),
                prize_source_key=cut_name,
            )

print(f"\n  Competitions: {S['comps_ins']} inserted, {S['comps_skip']} skipped")
print(f"  Classes:      {S['cls_ins']} inserted, {S['cls_skip']} skipped")
print(f"  Prizes:       {S['prizes']} inserted")

# ════════════════════════════════════════════════════════════════════════════
# STEP 4  Fabricate entries
# ════════════════════════════════════════════════════════════════════════════
print(f"\nSTEP 4 — Fabricating entries ({len(entries_todo)} classes)")
print("─" * 62)

r = sb.table("person").select("nationalid").like("nationalid", "000%").order("nationalid", desc=True).limit(1).execute()
max_nat = int(r.data[0]["nationalid"]) if r.data else 0
counter = max(COUNTER_START, max_nat + 1)
print(f"  Starting nationalid: {counter:09d}\n")

for cic_id, n_entries, cls_dt, comp_id, ranchid in entries_todo:
    ex_e    = sb.table("entry").select("entryid").eq("classincompid", cic_id).execute()
    already = len(ex_e.data)

    if already >= n_entries:
        print(f"  SKIP cic={cic_id} ({already}/{n_entries} exist)")
        counter += n_entries
        continue

    to_create = n_entries - already
    start_n   = counter + already

    for i in range(to_create):
        n   = start_n + i
        nat = f"{n:09d}"
        try:
            p = sb.table("person").insert({
                "nationalid": nat, "firstname": "רוכב", "lastname": f"היסטורי {n}",
            }).execute().data[0]
            pid = p["personid"]
            sb.table("federationmember").insert({"federationmemberid": pid, "hasvalidmembership": True}).execute()
            h = sb.table("horse").insert({"ranchid": ranchid, "horsename": f"סוס היסטורי {n}"}).execute().data[0]
            sb.table("systemuser").insert({
                "systemuserid": pid, "username": f"hist_{n}",
                "passwordhash": "placeholder", "passwordsalt": "placeholder", "isactive": False,
            }).execute()
            bill = sb.table("bill").insert({
                "paidbypersonid": pid, "amounttopay": 0,
                "dateopened": cls_dt.isoformat(), "competitionid": comp_id,
            }).execute().data[0]
            sr = sb.table("servicerequest").insert({
                "orderedbysystemuserid": pid, "horseid": h["horseid"],
                "riderfederationmemberid": pid, "billid": bill["billid"],
            }).execute().data[0]
            sb.table("entry").insert({
                "entryid": sr["srequestid"], "classincompid": cic_id, "entrystatus": "Active",
            }).execute()
            S["entries"] += 1
        except Exception as e:
            print(f"  ERROR entry n={nat} cic={cic_id}: {e}")
            S["errors"] += 1

        if S["entries"] > 0 and S["entries"] % 50 == 0:
            print(f"  ... {S['entries']} entries fabricated so far")

    counter += n_entries
    print(f"  cic={cic_id}: {to_create} entries fabricated")

# ════════════════════════════════════════════════════════════════════════════
# SUMMARY
# ════════════════════════════════════════════════════════════════════════════
print()
print("=" * 62)
print("INSERTION COMPLETE")
print("=" * 62)
print(f"  Competitions inserted  : {S['comps_ins']}")
print(f"  Classes inserted       : {S['cls_ins']}")
print(f"  Prizes inserted        : {S['prizes']}")
print(f"  Entries fabricated     : {S['entries']}")
print(f"  Rows skipped           : {S['rows_skip']}")
print(f"  New classtypes added   : {S['new_types']}")
print(f"  Errors                 : {S['errors']}")
print("=" * 62)
